import openpyxl
try:
    wb = openpyxl.load_workbook('D:\PROJECTS COLLEGE\Python_automation\Excel_to_mysql\data.xlsx')
except FileNotFoundError as err:
    print('No file found')
else:
    sheet = wb.active
    for column_name in sheet.iter_rows(min_row=1,max_row=1,values_only=True):
        print(column_name)
    for dtype in sheet.iter_rows(min_row=2,max_row=2,values_only=True):
        print(dtype)
    #print(value,dtype)

    dtype=list(dtype)
    
    for i in range(sheet.max_column):
        if type(dtype[i]) is str:
            dtype[i]='varchar'
        elif type(dtype[i]) is int:
            dtype[i]='int'
        elif type(dtype[i]) is float:
            dtype[i]='float'
        else:
            dtype[i]=dtype[i].strftime("%m/%d/%Y, %H:%M:%S")
            if '00:00:00' in dtype[i]:
                dtype[i]='date'
            else:
                dtype[i]='datetime'
    query=str()
    for i in range(len(column_name)):
        query+='{} {},'.format(column_name[i],dtype[i])
    print(query)

    #print('Final',column_name,dtype)



    row_count = sheet.max_row
    column_count = sheet.max_column
    #if row_count>0
    '''for row in sheet.iter_rows(min_row=1,max_row=5):  
        for cell in row:  
            print(type(cell.value), end=" ")  
        print()  
    
    '''

    