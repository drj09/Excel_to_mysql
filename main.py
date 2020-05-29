import mysql.connector
from mysql.connector import errorcode
import openpyxl

def sql_init():
    try:
        conn=mysql.connector.connect(user='root')
    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            print("Something is wrong with your user name or password")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            print("Database does not exist")
        else:
            print('---->',err)
        error()
    else:
        return conn

def sql_query(conn,query):
    mycursor = conn.cursor()
    mycursor.execute(query)
    try:
      myresult = mycursor.fetchall()
      return myresult
    except mysql.connector.Error as err:
        print('---->',err)
        return False
    finally:
        conn.commit()

#include path in this xl_file
def xl_file(path):
    print('##in xl_path')
    try:
        wb = openpyxl.load_workbook(path)
        #wb = openpyxl.load_workbook('D:\PROJECTS COLLEGE\Python_automation\Excel_to_mysql\data.xlsx')
    except Exception as err:
        print('---->',err)
        error()
    else:
        sheet = wb.active
        for column_name in sheet.iter_rows(min_row=1,max_row=1,values_only=True):
            print("Columns Found",column_name)
        for dtype in sheet.iter_rows(min_row=2,max_row=2,values_only=True):
            print('First Row',dtype)
        #print(value,dtype)

        dtype=list(dtype)
        column_name=list(column_name)
        
        for i in range(sheet.max_column):
            column_name[i]=column_name[i].replace('.','_')
            column_name[i]=column_name[i].replace(' ','_')            
            if type(dtype[i]) is str:
                dtype[i]='varchar(500)'
            elif type(dtype[i]) is int:
                dtype[i]='int'
            elif type(dtype[i]) is float:
                dtype[i]='float'
            else:
                dtype[i]='datetime'
                '''

                dtype[i]=dtype[i].strftime("%m/%d/%Y, %H:%M:%S")
                if '00:00:00' in dtype[i]:
                    dtype[i]='date'
                else:
                    dtype[i]='datetime'

                '''    
        tb_structure=str()
        for i in range(len(column_name)):
            tb_structure+='{} {},'.format(column_name[i],dtype[i])
        print('Data types of excel data --> ',tb_structure)
        return tb_structure[:-1],sheet
        

def check_database(conn,db_name):
    print('##in check database')
    db_list=sql_query(conn,"show databases like \'{}'".format(db_name).lower())
    try:
        if db_list:
            print('Database present', db_list)
        else:
            sql_query(conn,"create database {}".format(db_name).lower())
        sql_query(conn,f"use {db_name}")
    except Exception as err:
        print('---->',err)
        

def check_table(conn,tb_structure,tb_name):
    print('##in check_table')
    try:
        a=sql_query(conn,"create table {} ({})".format(tb_name,tb_structure))
        print(a)
    except Exception as err:
        print('---->',err)
        error()
    
    """tb_list=sql_query(conn,"show tables like \'{}'".format('ninja').lower())
    print(tb_list)
    if tb_list:
      print('Table present', tb_list)
      a=sql_query(conn,"show columns from ninja")
      print('a-->',a)
      print('tb_structure',tb_structure)
    else:
      sql_query(conn,"create table {}".format('Paa').lower())"""
    

def push_data(conn,sheet,tb_name):
    print('##in push_data')
    try:
        for dtype in sheet.iter_rows(min_row=2,values_only=True):
            a=str()
            qry=str()
            dtype=list(dtype)

            for i in range(len(dtype)):
                #print(dtype[i],type(dtype[i]),end='>>')
                if type(dtype[i]) is not int and type(dtype[i])is not float:
                    #print(dtype[i])
                    dtype[i]=f'"{dtype[i]}"'
                qry+=f'{dtype[i]}'+','
            qry=f'insert into {tb_name} values({qry[:-1]})'
            print(qry)
            a=sql_query(conn,qry)
            #print(a)
    except Exception as err:
        print('---->',err)
        error()


def error():
    print('Error occured')
    exit()



def main():
    conn=sql_init()
    tb_structure,sheet = xl_file(input("Enter full path of excel sheet"))

    print('---> main',tb_structure,sheet)
    
    check_database(conn,input('Enter database name').replace(' ',"_"))
    
    tb_name=input('ENTER Table name').replace(' ','_')

    
    check_table(conn,tb_structure,tb_name)
    
    push_data(conn,sheet,tb_name)
    
    
    #print(sql_query(conn,'insert into hehe values(11,255546,"Satya","15th","2020-12-21 00:00:00",252,32)'))
    
    #xl_path=input('Enter path of file')
    #xl_file(xl_path)
    '''
    db_name=input('Enter database name')
    check_database(db_name)
    tb_name=input('Enter table name')
    check_table(tb_name)
    '''
main()


